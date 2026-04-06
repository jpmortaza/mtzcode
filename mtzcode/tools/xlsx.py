"""Tools para ler e escrever planilhas .xlsx."""
from __future__ import annotations

from pathlib import Path

from pydantic import BaseModel, Field

from mtzcode.tools.base import Tool, ToolError


# ---------- Read ----------


class XlsxReadArgs(BaseModel):
    path: str = Field(..., description="Caminho do .xlsx a ler.")
    sheet: str | None = Field(None, description="Nome da sheet. Se None, usa a primeira.")
    max_rows: int = Field(50, description="Máximo de linhas a retornar.")


class XlsxReadTool(Tool):
    name = "xlsx_read"
    destructive = False
    description = (
        "Lê uma planilha .xlsx e devolve as primeiras linhas como tabela markdown. "
        "Lista também as sheets disponíveis."
    )
    Args = XlsxReadArgs

    def run(self, args: XlsxReadArgs) -> str:  # type: ignore[override]
        # Import lazy do openpyxl.
        try:
            from openpyxl import load_workbook  # type: ignore
        except ImportError as exc:
            raise ToolError(
                "dependência ausente: instale `openpyxl` (pip install openpyxl)"
            ) from exc

        p = Path(args.path).expanduser()
        if not p.exists():
            raise ToolError(f"arquivo não encontrado: {p}")

        try:
            wb = load_workbook(str(p), data_only=True, read_only=True)
        except Exception as exc:  # noqa: BLE001
            raise ToolError(f"erro abrindo {p}: {exc}") from exc

        sheets = wb.sheetnames
        nome = args.sheet if args.sheet else (sheets[0] if sheets else None)
        if nome is None:
            return "(workbook sem sheets)"
        if nome not in sheets:
            raise ToolError(f"sheet `{nome}` não existe. disponíveis: {sheets}")

        ws = wb[nome]

        # Junta as primeiras max_rows linhas em uma matriz de strings.
        rows: list[list[str]] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= args.max_rows:
                break
            rows.append(["" if v is None else str(v) for v in row])

        header_lines = [
            f"sheets: {sheets}",
            f"sheet ativa: {nome}",
            f"linhas exibidas: {len(rows)} (limite {args.max_rows})",
            "",
        ]

        if not rows:
            return "\n".join(header_lines + ["(sheet vazia)"])

        ncols = max(len(r) for r in rows)
        # Normaliza largura das linhas.
        rows = [r + [""] * (ncols - len(r)) for r in rows]

        header = rows[0]
        md = ["| " + " | ".join(header) + " |", "| " + " | ".join("---" for _ in header) + " |"]
        for r in rows[1:]:
            md.append("| " + " | ".join(r) + " |")

        return "\n".join(header_lines + md)


# ---------- Write ----------


class XlsxWriteArgs(BaseModel):
    path: str = Field(..., description="Caminho de saída do .xlsx.")
    data: list[list[str]] = Field(..., description="Matriz de linhas/colunas a escrever.")
    sheet_name: str = Field("Sheet1", description="Nome da sheet a criar.")


class XlsxWriteTool(Tool):
    name = "xlsx_write"
    destructive = True
    description = (
        "Cria um arquivo .xlsx a partir de uma matriz de strings (lista de listas). "
        "A primeira linha geralmente é o header."
    )
    Args = XlsxWriteArgs

    def run(self, args: XlsxWriteArgs) -> str:  # type: ignore[override]
        # Import lazy.
        try:
            from openpyxl import Workbook  # type: ignore
        except ImportError as exc:
            raise ToolError(
                "dependência ausente: instale `openpyxl` (pip install openpyxl)"
            ) from exc

        p = Path(args.path).expanduser()
        p.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = args.sheet_name

        for row in args.data:
            ws.append(list(row))

        try:
            wb.save(str(p))
        except OSError as exc:
            raise ToolError(f"erro salvando {p}: {exc}") from exc

        return f"salvo: {len(args.data)} linhas em {p}"
